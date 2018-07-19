# -*-coding: utf-8-*-
import requests
from openpyxl import Workbook
import datetime
import sys
import pandas as pd


HEADERS    = {'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_10_1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/39.0.2171.95 Safari/537.36'}
COINS      = ['BTC','NEO']
CURRENCY   = ['KRW','BTC']
TIME_TYPE  = ['days','minutes']
MNTS_TYPE  = [60, 240]
TIME_COUNT = 10000000


def printException():
    message = []
    for e in sys.exc_info():
        message.append(str(e))
    message = '\n'.join(message)
    print(message)
    return message


def makeUpbitURL(coinName, timeType, timeBucket, timeCount, toDateTime, currency) :
    url = 'https://crix-api-endpoint.upbit.com/v1/crix/candles/%s/%s?code=CRIX.UPBIT.%s-%s&count=%d&to=%s'% (timeType, str(timeBucket), currency, coinName, timeCount, toDateTime)
    print('url : ', url)
    return url

def writeExcelUpbitResultByOpenpyxl(coinName, currency, dataFrame, timeType, timeBucket) :
    print('Write Excel ...')
    print(' ')

    dataFrame = dataFrame.sort_values(by=['candleDateTime'])

    wb = Workbook()

    try :
        ws = wb.active

        ws['A1'] = "Time"
        ws['B1'] = "TimeKst"
        ws['C1'] = 'Open'
        ws['D1'] = 'High'
        ws['E1'] = 'Low'
        ws['F1'] = 'Close'
        ws['G1'] = 'Volume'

        i = 0
        for index, df in dataFrame.iterrows() :
            ws.cell(row=i + 2, column=1).value = df['candleDateTime']
            ws.cell(row=i + 2, column=2).value = df['candleDateTimeKst']
            ws.cell(row=i + 2, column=3).value = df['openingPrice']
            ws.cell(row=i + 2, column=4).value = df['highPrice']
            ws.cell(row=i + 2, column=5).value = df['lowPrice']
            ws.cell(row=i + 2, column=6).value = df['tradePrice']
            ws.cell(row=i + 2, column=7).value = df['candleAccTradeVolume']

            i = i + 1

        saveStr = 'Data/Upbit_%s%s_%s_%s.xlsx' % (coinName, currency, timeType, timeBucket)

        wb.save(saveStr)

    except:
        printException()
    finally:
        wb.close()

def printResult(data) :
    code = data[0]['code']
    print(code)
    print("==========================================")
    print("   date    open high low final    vol")
    print("==========================================")

    for i in range(len(data)):
        date = data[i]['candleDateTimeKst']
        onlyDate = date.split('T')  # 날짜정보와 시간정보 분리
        print(onlyDate[0], "%d" % data[i]['openingPrice'], "%f" % data[i]['highPrice'], "%f" % data[i]['lowPrice'],
              "%f" % data[i]['tradePrice'], "%f" % data[i]['candleAccTradeVolume'], "%d" % data[i]['timestamp'], str(datetime.datetime.fromtimestamp(data[i]['timestamp']/1000)))

    print("==========================================")



if __name__ == "__main__":

    print('Start Extract.....')
    print(' ')

    for coin in COINS :

        for crncy in CURRENCY :
            if coin == crncy : continue

            for timeType in TIME_TYPE :

                if timeType == 'days' :

                    print('Extract : ' + coin + '-' + crncy + ' ' + timeType + '...')

                    toDateTime = ''
                    total_df = pd.DataFrame()

                    while True :

                        url = makeUpbitURL(coin, timeType, '', TIME_COUNT, toDateTime, crncy)

                        try:
                            html = requests.get(url, headers=HEADERS)
                        except requests.exceptions.HTTPError as err:
                            print(err)
                            exit(1)

                        df = pd.DataFrame.from_records(html.json())

                        if len(df) == 0: break

                        total_df = total_df.append(df)

                        orgToDateTime = str(df.iloc[len(df) - 1]['candleDateTime'])
                        splitedToDateTime = orgToDateTime.split('+')
                        toDateTime = splitedToDateTime[0][:10] + ' ' + splitedToDateTime[0][11:]

                        print('Extract : ' + coin + '-' + crncy + ' ' + timeType + ' ---> End datetime : ' + toDateTime)

                    # printResult(df)
                    writeExcelUpbitResultByOpenpyxl(coin, crncy, total_df, timeType, 1)

                elif timeType == 'minutes' :

                    for timeBucket in MNTS_TYPE :

                        print('Extract : ' + coin + '-' + crncy + ' ' + timeType + ' ' + str(timeBucket) + '...')

                        toDateTime = ''
                        total_df = pd.DataFrame()

                        while True :

                            url = makeUpbitURL(coin, timeType, timeBucket, TIME_COUNT, toDateTime, crncy)

                            try:
                                html = requests.get(url, headers=HEADERS)
                            except requests.exceptions.HTTPError as err:
                                print(err)
                                exit(1)

                            df = pd.DataFrame.from_records(html.json())

                            if len(df) == 0: break

                            total_df = total_df.append(df)

                            orgToDateTime = str(df.iloc[len(df) - 1]['candleDateTime'])
                            splitedToDateTime = orgToDateTime.split('+')
                            toDateTime = splitedToDateTime[0][:10] + ' ' + splitedToDateTime[0][11:]

                            print('Extract : ' + coin + '-' + crncy + ' ' + timeType + ' ' + str(timeBucket) + ' ---> End datetime : ' + toDateTime)

                        # printResult(df)
                        writeExcelUpbitResultByOpenpyxl(coin, crncy, total_df, timeType, timeBucket)